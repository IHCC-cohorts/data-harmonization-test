#!/usr/bin/env python3

import cgi
import cogs
import csv
import os
import re
import shutil
import subprocess
import sys
import urllib.parse
import json

from openpyxl import load_workbook
from cogs.helpers import get_sheet_url

output_format = "text"
prefixes = {}

email_pattern = re.compile("^\S+@\S+$")
cohort_pattern = re.compile("^[A-Z0-9]+$")
file_pattern = re.compile("\"(/.+/ring-multipart-\d+.tmp)\"")
sheet_pattern = re.compile("^\w+$")


def main():
    output = ["h1", "Unhandled input"]
    build = "../build"

    # Read arguments from STDIN as a query string
    args = {}
    fields = None
    if os.environ["REQUEST_METHOD"] == "POST":
        fields = cgi.FieldStorage()
        args = {}
        for key in fields.keys():
            args[key] = fields[key].value
    elif "QUERY_STRING" in os.environ:
        args = dict(urllib.parse.parse_qsl(os.environ["QUERY_STRING"]))
    # TODO: read command-line arguments

    if not "action" in args or not args["action"]:
        output = ["h1", "Error: please specify an action"]
        # output.append(["p", str(args)])
        # output.append(["p", "STDIN: ", str(sys.stdin.read())])
        return render_output(output)

    action = args["action"]

    # If COGS has already run, redirect to the Google Sheet
    if action == "open":
        return open_sheet(args)

    # Create a new cohort entry: show empty form.
    elif action == "create":
        if os.path.exists("../.cogs/config.tsv"):
            return open_sheet(args)
        return build_form(args)

    # Validate the submitted form.
    elif action == "upload":
        if os.path.exists("../.cogs/config.tsv"):
            return open_sheet(args)

        valid = {}
        invalid = {}
        wb = None

        for name in ["admin_google_id", "submitter_google_id"]:
            if not name in args:
                invalid[name] = "This is a required field"
            else:
                args[name] = args[name].strip()
                if not email_pattern.match(args[name]):
                    invalid[name] = "Must be a valid email address"
                else:
                    valid[name] = True

        name = "cohort_id"
        if not name in args:
            invalid[name] = "This is a required field"
        else:
            args[name] = args[name].strip()
            path = f"../templates/{args[name]}.tsv".lower()
            if not cohort_pattern.match(args[name]):
                invalid[name] = "Cohort ID must contain only uppercase letters and numbers"
            elif os.path.isfile(path):
                invalid[name] = "Cohort ID must not conflict with existing ID"
            else:
                valid[name] = True

        name = "upload_template"
        if not name in args:
            invalid[name] = "This is a required field"
        else:
            try:
                wb = load_workbook(fields[name].file)
                if "Instructions" not in wb.sheetnames:
                    invalid[name] = "Instructions sheet is required"
                elif "Metadata" not in wb.sheetnames:
                    invalid[name] = "Metadata sheet is required"
                elif "Terminology" not in wb.sheetnames:
                    invalid[name] = "Terminology sheet is required"
                else:
                    valid[name] = True
            except Exception as e:
                invalid[name] = f"Not a valid Excel file: {e}"

        if invalid:
            args["valid"] = valid
            args["invalid"] = invalid
            return build_form(args)

        os.makedirs(build, exist_ok=True)
        with open(build + "/submitter_google_id", "w") as f:
            f.write(args["submitter_google_id"])

        cohort_id = args["cohort_id"]
        instructions = f"build/instructions.tsv".lower()
        # TODO: These are the paths that I want once we fix https://github.com/ontodev/cogs/issues/59
        # metadata = f"metadata/{cohort_id}.tsv".lower()
        # terminology = f"templates/{cohort_id}.tsv".lower()
        metadata = f"build/metadata.tsv".lower()
        terminology = f"build/terminology.tsv".lower()

        save_sheet(wb["Instructions"], "../" + instructions)
        save_sheet(wb["Metadata"], "../" + metadata)
        save_sheet(wb["Terminology"], "../" + terminology)

        # TODO: These would be better as Python calls.
        os.chdir("..")
        cogs.init(
            f"IHCC Data Harmonization: {cohort_id}",
            user=args["admin_google_id"],
            role="writer",
        )
        cogs.add(instructions, title="Instructions")
        cogs.add(metadata, title="Metadata")
        cogs.add(terminology, title="Terminology", freeze_row=1)
        cogs.push()
        link = get_sheet_url()

        output = [
            "div",
            ["p", f"Google Sheet created and shared with '{args['admin_google_id']}'."],
            [
                "ul",
                ["li", ["a", {"href": link, "target": "_blank"}, "Open Google Sheet"]],
            ],
        ]
        return render_output(output)

    if action == "share":
        submitter_google_id = None
        if not os.path.exists("../.cogs/config.tsv"):
            output = [
                "div",
                ["h1", "Error: No Google Sheet has been configured"],
                ["p", ["a", {"href": "workflow.py?action=create"}, "Upload cohort data"]],
            ]
            return render_output(output)

        if "submitter_google_id" in args:
            submitter_google_id = args["submitter_google_id"]
        else:
            try:
                with open(build + "/submitter_google_id") as f:
                    submitter_google_id = f.read().strip()
            except:
                pass

        invalid = {}
        name = "submitter_google_id"
        if submitter_google_id and not email_pattern.match(submitter_google_id):
            invalid[name] = "Must be a valid email address"
        if not submitter_google_id or invalid:
            args["invalid"] = invalid
            output = [
                "form",
                {"action": "workflow.py"},
                ["p", "Share Google Sheet with submitter:"],
                ["input", {"type": "hidden", "name": "action", "value": "share"}],
                build_input(args, "Submitter Google ID"),
                build_input(args, "Submit", input_type="submit"),
            ]
            return render_output(output)

        os.chdir("..")
        cogs.share(submitter_google_id, "writer")
        link = get_sheet_url()
        output = [
            "div",
            ["h1", "Google Sheet Shared"],
            ["p", submitter_google_id],
            ["li", ["a", {"href": link, "target": "_blank"}, "Open Google Sheet"]],
        ]
        return render_output(output)


def save_sheet(ws, path):
    with open(path, "w") as fh:
        writer = csv.writer(fh, delimiter="\t", lineterminator="\n")
        for row in ws.values:
            writer.writerow(row)


def open_sheet(args):
    if os.path.exists("../.cogs/config.tsv"):
        link = get_sheet_url()
        output = [
            "ul",
            ["li", ["a", {"href": link, "target": "_blank"}, "Open Google Sheet"]],
            ["script", {"type": "text/javascript"}, f"window.open('{link}', '_blank');"]
            # ["meta", {"http-equiv": "refresh", "content": f"0; URL=../.."}]
        ]
    else:
        output = [
            "div",
            ["h1", "Error: No Google Sheet has been configured"],
            ["p", ["a", {"href": "workflow.py?action=create"}, "Upload cohort data"]],
        ]
    return render_output(output)


def build_form(args):
    output = [
        "form",
        {"action": "workflow.py", "method": "POST", "enctype": "multipart/form-data"},
        ["p", "Submit a new cohort:"],
        # ["p", str(args)],
        ["input", {"type": "hidden", "name": "action", "value": "upload"}],
        build_input(args, "Admin Google ID"),
        build_input(args, "Submitter Google ID"),
        build_input(args, "Cohort ID"),
        build_input(args, "Upload Template", input_type="file"),
        build_input(args, "Submit", input_type="submit"),
    ]
    return render_output(output)


def build_input(args, label, input_type="text"):
    output = ["div", {"class": "form-group row"}]
    name = label.lower().replace(" ", "_")
    value = args[name] if name in args else ""
    left = "col-md-3"
    right = "col-md-9"
    control = [right, "form-control"]
    if "valid" in args and name in args["valid"]:
        control.append("is-valid")
    elif "invalid" in args and name in args["invalid"]:
        control.append("is-invalid")
    label_classes = " ".join([left, "col-form-label"])
    control_classes = " ".join(control)

    if input_type == "textarea":
        output.append(["label", {"for": name, "class": label_classes}, label])
        output.append(["textarea", {"class": control_classes, "name": name}, value])
    elif input_type == "file":
        output.append(["label", {"for": name, "class": label_classes}, label])
        control_classes = control_classes.replace("form-control", "form-control-file")
        output.append(
            ["input", {"type": "file", "class": control_classes, "name": name, "value": label}]
        )
    elif input_type == "submit":
        output.append(
            ["input", {"type": "submit", "class": "btn btn-primary", "name": name, "value": label}]
        )
    else:
        output.append(["label", {"for": name, "class": label_classes}, label])
        output.append(
            ["input", {"type": "text", "class": control_classes, "name": name, "value": value}]
        )

    if "valid" in args and name in args["valid"] and isinstance(args["valid"][name], str):
        output.append(["div", {"class": left}])
        output.append(["div", {"class": right + " valid-feedback"}, args["valid"][name]])
    if "invalid" in args and name in args["invalid"] and isinstance(args["invalid"][name], str):
        output.append(["div", {"class": left}])
        output.append(["div", {"class": right + " invalid-feedback"}, args["invalid"][name]])

    return output


def render_output(output):
    # Render output
    if os.environ.get("GATEWAY_INTERFACE") == "CGI/1.1":
        print("Content-Type: text/html")
        print("")
        print(render_html(prefixes, output))
    else:
        print(render_text(output))


def render_html(prefixes, element, depth=0):
    """Render hiccup-style HTML vector as HTML."""
    indent = "  " * depth
    if not isinstance(element, list):
        raise Exception(f"Element is not a list: {element}")
    if len(element) == 0:
        raise Exception(f"Element is an empty list")
    tag = element.pop(0)
    if not isinstance(tag, str):
        raise Exception(f"Tag '{tag}' is not a string in '{element}'")
    output = f"{indent}<{tag}"

    if len(element) > 0 and isinstance(element[0], dict):
        attrs = element.pop(0)
        if tag == "a" and "href" not in attrs and "resource" in attrs:
            attrs["href"] = curie2href(attrs["resource"])
        for key, value in attrs.items():
            if key in ["checked"]:
                if value:
                    output += f" {key}"
            else:
                output += f' {key}="{value}"'

    if tag in ["meta", "link"]:
        output += "/>"
        return output
    output += ">"
    spacing = ""
    if len(element) > 0:
        for child in element:
            if isinstance(child, str):
                output += child
            elif isinstance(child, list):
                try:
                    output += "\n" + render_html(prefixes, child, depth=depth + 1)
                    spacing = f"\n{indent}"
                except Exception as e:
                    raise Exception(f"Bad child in '{element}'", e)
            else:
                raise Exception(f"Bad type for child '{child}' in '{element}'")
    output += f"{spacing}</{tag}>"
    return output


def render_text(element):
    """Render hiccup-style HTML vector as text."""
    if not isinstance(element, list):
        raise Exception(f"Element is not a list: {element}")
    if len(element) == 0:
        raise Exception(f"Element is an empty list")
    tag = element.pop(0)
    output = ""
    if len(element) > 0:
        for child in element:
            if isinstance(child, str):
                output += child
            elif isinstance(child, list):
                try:
                    output += render_text(child)
                except Exception as e:
                    raise Exception(f"Bad child in '{element}'", e)
            elif isinstance(child, dict):
                pass
            else:
                raise Exception(f"Bad type for child '{child}' in '{element}'")
    return output


if __name__ == "__main__":
    main()
